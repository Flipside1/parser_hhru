import time
import random

from bs4 import BeautifulSoup
import requests

from selenium import webdriver
from selenium.webdriver.common.by import By

from openpyxl import load_workbook


driver = webdriver.Firefox()

fn = 'data_table.xlsx'
wb = load_workbook(fn)  # открывает саму таблицу
ws = wb['data']  # открывает определенный лист в таблице
ws['A1'] = 'НАЗВАНИЕ ВАКАНСИИ'
ws['B1'] = 'ССЫЛКА НА ВАКАНСИЮ'

class Search_vacancies:

    def __init__(self):
        self.pages = None
        self.counter = 1  # счетчик, нужен для вписывания элемента в определенную строку таблицы

    def parsing_names(self):
        """Ищет названия вакансий и ссылки на них
        После этого сохраняет в таблицу Excel"""

        search_name = ['django', 'python', 'selenium', 'sql', 'backend']  # по каким тегам ищутся вакансии
        for url in search_name:
            driver.get(f'https://kirov.hh.ru/search/vacancy?text={url}')  # вписывает в поле слово вакансии
            time.sleep(random.randrange(5, 10))

            self.amount_pages()  # считает количество страниц

            for page in range(self.pages):  # переходит с одной страницы на другую
                driver.get(
                    f'https://kirov.hh.ru/search/vacancy?text=django&from=suggest_post&salary=&clusters=true&ored_clusters=true&enable_snippets=true&page={page}&hhtmFrom=vacancy_search_list')

                driver.find_elements(By.CLASS_NAME, 'bloko-link')  # ищет класс, в котором содержатся названия вакансий
                elements = driver.find_elements(By.TAG_NAME, 'a')  # ищет элементы с тегом "а"

                for element in elements:  # перебирает все элементы, которые спарсились
                    if ('Python' in element.text) or ('Django' in element.text):  # если в названии вакансии есть определенные слова
                        self.counter += 1  # счетчик +1 каждый раз
                        name = element.text  # выводит текст названия вакансии
                        url = element.get_attribute('href')  # находит ссылки на вакансии
                        ws[f'A{self.counter}'] = name  # вписывает названия вакансий в таблицу
                        ws[f'B{self.counter}'] = url  # вписывает ссылки на вакансии в таблицу
                        wb.save(fn)  # сохраняет таблицу

        driver.close()  # закрывает браузер

    def amount_pages(self):
        """Ищет количество страниц на сайте"""
        driver.find_element(By.CLASS_NAME, 'pager')  # ищет элемент класса (чтобы посчитать количество страниц)
        # записывает в переменную количество страниц на сайте
        pages = driver.find_element(By.CSS_SELECTOR,
                                    '#HH-React-Root > div > div.HH-MainContent.HH-Supernova-MainContent > div.main-content > div > div:nth-child(3) > div.sticky-sidebar-and-content--NmOyAQ7IxIOkgRiBRSEg > div.bloko-column.bloko-column_xs-4.bloko-column_s-8.bloko-column_m-9.bloko-column_l-13 > div > div.bloko-gap.bloko-gap_top > div > span:nth-child(6) > span.pager-item-not-in-short-range > a > span')
        self.pages = int(pages.text)  # преобразовывает текст в числовой формат
        print(self.pages)

    def transition_to_page(self):
        """Переходит на каждую страницу на сайте"""
        for i in range(self.pages):  # переходит с одной ссылки (страницы) на другую
            driver.get(f'https://kirov.hh.ru/search/vacancy?text=django&from=suggest_post&salary=&clusters=true&ored_clusters=true&enable_snippets=true&page={i}&hhtmFrom=vacancy_search_list')


search = Search_vacancies()

if __name__ == '__main__':
    search.parsing_names()
    wb.close()  # закрывает таблицу
else:
    print('ERROR')
